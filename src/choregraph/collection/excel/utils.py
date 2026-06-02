# SPDX-FileCopyrightText: 2026 NVEIL SAS
# SPDX-FileContributor: Clément Baraille
# SPDX-FileContributor: Pierre Jacquet
# SPDX-FileContributor: Guillaume Franque
# SPDX-License-Identifier: AGPL-3.0-or-later

import logging
import os
import re
import tempfile
import unicodedata
from collections import Counter
from pathlib import Path
from typing import List, Optional, Dict, Any

import pandas as pd
from openpyxl.utils import range_boundaries, get_column_letter

from .cartograph import TableStructure, SpreadsheetMap
from ...security import safe_path

logger = logging.getLogger(__name__)


_temp_files: List[str] = []


def _restore_xls_merges(source_xls: str, converted_xlsx: str):
    """Restore merged-cell values lost during pyexcel .xls → .xlsx conversion.

    pyexcel strips merge info: non-origin cells in a merged range become empty
    (or retain stale phantom values). This reads the original merge ranges via
    xlrd, then overwrites every cell in each range with the top-left value in
    the converted .xlsx file.
    """
    try:
        import xlrd
    except ImportError:
        logger.warning("xlrd not installed — skipping merge restoration for .xls file")
        return

    try:
        xls_wb = xlrd.open_workbook(source_xls, formatting_info=True)
    except Exception:
        # formatting_info not supported for all .xls variants; degrade gracefully
        logger.warning("Could not read .xls formatting info — skipping merge restoration")
        return

    import openpyxl as _openpyxl
    xlsx_wb = _openpyxl.load_workbook(converted_xlsx)

    total_filled = 0
    for xls_sheet in xls_wb.sheets():
        if xls_sheet.name not in xlsx_wb.sheetnames:
            continue
        xlsx_ws = xlsx_wb[xls_sheet.name]
        for rlo, rhi, clo, chi in xls_sheet.merged_cells:
            # xlrd: rlo/clo 0-based, rhi/chi exclusive
            top_left_value = xls_sheet.cell_value(rlo, clo)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    # openpyxl is 1-based
                    xlsx_ws.cell(row=r + 1, column=c + 1).value = top_left_value
                    total_filled += 1

    if total_filled:
        xlsx_wb.save(converted_xlsx)
        logger.info("Restored %d merged cells from original .xls", total_filled)
    xlsx_wb.close()


def _restore_ods_merges(source_ods: str, converted_xlsx: str):
    """Restore merged-cell values lost during pyexcel .ods → .xlsx conversion.

    Same problem as .xls: pyexcel strips merge info. This reads the original
    merge ranges via odfpy, then overwrites every cell in each range with the
    top-left value in the converted .xlsx file.
    """
    try:
        from odf.opendocument import load as odf_load
        from odf.table import Table, TableRow, TableCell
        from odf.namespaces import TABLENS
    except ImportError:
        logger.warning("odfpy not installed — skipping merge restoration for .ods file")
        return

    try:
        ods_doc = odf_load(source_ods)
    except Exception:
        logger.warning("Could not read .ods file — skipping merge restoration")
        return

    import openpyxl as _openpyxl
    xlsx_wb = _openpyxl.load_workbook(converted_xlsx)

    total_filled = 0
    for table in ods_doc.spreadsheet.getElementsByType(Table):
        sheet_name = table.getAttribute("name")
        if sheet_name not in xlsx_wb.sheetnames:
            continue
        xlsx_ws = xlsx_wb[sheet_name]

        row_idx = 0
        for row_elem in table.getElementsByType(TableRow):
            row_repeat = int(row_elem.getAttrNS(TABLENS, "number-rows-repeated") or 1)

            col_idx = 0
            for cell_elem in row_elem.getElementsByType(TableCell):
                col_span = int(cell_elem.getAttrNS(TABLENS, "number-columns-spanned") or 1)
                row_span = int(cell_elem.getAttrNS(TABLENS, "number-rows-spanned") or 1)
                col_repeat = int(cell_elem.getAttrNS(TABLENS, "number-columns-repeated") or 1)

                if col_span > 1 or row_span > 1:
                    # Read top-left value from the converted xlsx (pyexcel got it right)
                    top_left_value = xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1).value
                    for r in range(row_idx, row_idx + row_span):
                        for c in range(col_idx, col_idx + col_span):
                            xlsx_ws.cell(row=r + 1, column=c + 1).value = top_left_value
                            total_filled += 1

                col_idx += col_repeat * col_span

            row_idx += row_repeat

    if total_filled:
        xlsx_wb.save(converted_xlsx)
        logger.info("Restored %d merged cells from original .ods", total_filled)
    xlsx_wb.close()


def get_compatible_excel_path(source_path: str) -> str:
    """
    Assure que le fichier est au format .xlsx pour openpyxl.
    Si c'est un .xls ou .ods, on le convertit via pyexcel dans un fichier temporaire.
    For .xls/.ods files, merged-cell values are restored after conversion (pyexcel
    strips merge info, leaving phantom/empty values in non-origin cells).
    Temp files are tracked and cleaned up by cleanup_temp_files().
    """
    path = Path(source_path)
    extension = path.suffix.lower()

    if extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        return source_path

    logger.info("Converting %s format to .xlsx via pyexcel", extension)

    try:
        import pyexcel
    except ImportError:
        raise RuntimeError(
            f"pyexcel is required to convert {extension} files. "
            "Install it with: pip install pyexcel pyexcel-xls pyexcel-ods3"
        )

    target_path = str(path.with_suffix(".xlsx"))

    try:
        pyexcel.save_book_as(file_name=source_path, dest_file_name=target_path)
    except Exception as e:
        raise RuntimeError(f"Failed to convert via pyexcel: {e}")

    # Restore merged-cell values that pyexcel lost during conversion
    if extension == '.xls':
        _restore_xls_merges(source_path, target_path)
    elif extension == '.ods':
        _restore_ods_merges(source_path, target_path)

    try:
        os.remove(safe_path(source_path))
    except OSError as e:
        logger.warning("Could not remove original file %s: %s", source_path, e)
        
    return target_path


def cleanup_temp_files():
    """Remove any temporary converted files created by get_compatible_excel_path."""
    for fpath in _temp_files:
        try:
            os.remove(fpath)
        except OSError:
            pass
    _temp_files.clear()


def _slugify(text: str, max_length: int = 80) -> str:
    """Convert a human-readable label into a filesystem-safe snake_case slug.

    NFKD normalize → strip combining marks → lowercase → replace
    non-alphanumeric with ``_`` → collapse consecutive ``_`` → truncate.
    Falls back to ``"table"`` if the result is empty.
    """
    # Normalize unicode and strip combining characters (accents)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    text = text[:max_length].rstrip("_")
    return text or "table"


def _normalize_name(name: str) -> str:
    """Normalize a column name for comparison: NFKD, strip accents, lowercase, strip whitespace."""
    name = unicodedata.normalize("NFKD", str(name))
    name = "".join(ch for ch in name if not unicodedata.combining(ch))
    return name.lower().strip()


def _align_column_names(results: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """Unify column names across DataFrames that differ only by accent/case/whitespace.

    For each group of column names that share the same normalized form,
    the most frequent original spelling is chosen as canonical.  All
    DataFrames are then renamed in-place.
    """
    # Build normalized_form -> Counter({original_name: occurrence_count})
    norm_to_originals: Dict[str, Counter] = {}
    for df in results.values():
        for col in df.columns:
            norm = _normalize_name(col)
            if norm not in norm_to_originals:
                norm_to_originals[norm] = Counter()
            norm_to_originals[norm][col] += 1

    # Build rename map: original_variant -> canonical (most frequent)
    rename_map: Dict[str, str] = {}
    for norm, counter in norm_to_originals.items():
        if len(counter) <= 1:
            continue  # No variants to unify
        canonical = counter.most_common(1)[0][0]
        for variant in counter:
            if variant != canonical:
                rename_map[variant] = canonical

    if not rename_map:
        return results

    logger.info("Aligning column names: %s", rename_map)
    return {
        key: df.rename(columns=rename_map)
        for key, df in results.items()
    }


# =============================================================================
# FONCTIONS UTILITAIRES
# =============================================================================

def extract_sub_table(df_initial: pd.DataFrame, excel_range: str) -> pd.DataFrame:
    """
    Extrait un sous-tableau d'un DataFrame à partir d'une plage Excel.
    Validates the range against actual DataFrame dimensions.
    """
    min_col, min_row, max_col, max_row = range_boundaries(excel_range)

    # R2: Validate that the range makes sense
    if min_row > max_row or min_col > max_col:
        raise ValueError(f"Invalid range '{excel_range}': inverted boundaries")

    df_rows, df_cols = df_initial.shape
    if min_row < 1 or min_col < 1:
        raise ValueError(f"Invalid range '{excel_range}': coordinates must be >= 1")

    # Clamp to actual dimensions with a warning
    clamped_max_row = min(max_row, df_rows)
    clamped_max_col = min(max_col, df_cols)
    if clamped_max_row < max_row or clamped_max_col < max_col:
        logger.warning(
            "Range '%s' exceeds sheet dimensions (%d rows x %d cols), clamped to row %d col %d",
            excel_range, df_rows, df_cols, clamped_max_row, clamped_max_col
        )

    sub_df = df_initial.iloc[min_row - 1 : clamped_max_row, min_col - 1 : clamped_max_col]
    return sub_df.copy().reset_index(drop=True)


def df_to_markdown(df: pd.DataFrame, max_rows: int = 15) -> str:
    """
    Convertit un DataFrame en table Markdown pour les prompts LLM.
    
    Args:
        df: DataFrame à convertir
        max_rows: Nombre max de lignes à afficher
        
    Returns:
        String Markdown de la table
    """
    # Limiter le nombre de lignes
    if len(df) > max_rows:
        df_display = pd.concat([df.head(max_rows - 2), df.tail(2)])
    else:
        df_display = df
    
    # Construire le header
    header_cells = [str(col).replace("|", "\\|") for col in df_display.columns]
    markdown = "| " + " | ".join(header_cells) + " |\n"
    markdown += "|" + "|".join(["---"] * len(header_cells)) + "|\n"
    
    # Construire les lignes
    for idx, (_, row) in enumerate(df_display.iterrows()):
        if len(df) > max_rows and idx == max_rows - 2:
            markdown += "| ... | " + " | ".join(["..."] * (len(header_cells) - 1)) + " |\n"
        values = [
            str(value).replace("|", "\\|")[:30] if pd.notna(value) else ""
            for value in row.values
        ]
        markdown += "| " + " | ".join(values) + " |\n"
    
    return markdown


def get_llm_friendly_columns(sub_df: pd.DataFrame, head: int = 8, tail: int = 3) -> str:
    """
    Génère une description des colonnes pour le prompt LLM.
    
    Affiche seulement head + tail colonnes si trop nombreuses.
    Inclut un exemple de valeur pour chaque colonne.
    """
    columns = list(sub_df.columns)
    total = len(columns)
    
    if total <= head + tail:
        indices_to_show = list(range(total))
        skip_info = None
    else:
        indices_to_show = list(range(head)) + list(range(total - tail, total))
        skip_info = (head, total - tail - 1, total - head - tail)
    
    lines = []
    for i, idx in enumerate(indices_to_show):
        if skip_info and i == head:
            lines.append(f"  ... ({skip_info[2]} colonnes omises: index {skip_info[0]} → {skip_info[1]}) ...")
        
        col = columns[idx]
        first_val = sub_df[col].dropna().iloc[0] if not sub_df[col].dropna().empty else "vide"
        first_val_str = str(first_val)
        if len(first_val_str) > 30:
            first_val_str = first_val_str[:30] + "..."
        lines.append(f"- [{idx}] Column '{col}' (ex: \"{first_val_str}\")")
    
    header = f"📊 {total} colonnes au total:\n"
    return header + "\n".join(lines)


def print_table_structure(table: 'TableStructure') -> str:
    """Affichage formaté d'une TableStructure."""
    pretty = f"📊 [{table.table_id}] {table.label}\n"
    pretty += f"   Sheet: {table.sheet_name}\n"
    pretty += f"   Type: {table.type} | Orientation: {table.orientation}\n"
    pretty += f"   Pattern: {table.pattern_type} | Header Complexity: {table.header_complexity}\n"
    pretty += f"   Full Range: {table.full_range.range}\n"
    pretty += f"   Has Headers: {table.has_headers}\n"
    if table.header_range:
        pretty += f"   Header Range: {table.header_range.range}\n"
    else:
        pretty += f"   Header Range: None (no headers)\n"
    pretty += f"   Data Range: {table.data_range.range}\n"
    if table.identity_columns:
        pretty += f"   Identity Columns: {', '.join(table.identity_columns)}\n"
    return pretty


def print_spreadsheet_map(sm: 'SpreadsheetMap') -> str:
    """Affichage formaté d'une SpreadsheetMap."""
    pretty = "=" * 60 + "\n"
    pretty += f"Spreadsheet Map ({len(sm.detected_tables)} table(s))\n"
    pretty += "=" * 60 + "\n"

    for table in sm.detected_tables:
        pretty += "\n" + print_table_structure(table)

    if sm.sheet_global_context:
        pretty += "\nGlobal Context:\n"
        for context in sm.sheet_global_context:
            pretty += f"   - {context}\n"

    pretty += "\n" + "=" * 60 + "\n"
    return pretty


def print_region_mapping(mapping: 'RegionMapping') -> str:
    """Affichage formaté d'un RegionMapping."""
    pretty = f"🗺️ RegionMapping for {mapping.table_id}\n"
    pretty += f"   Pattern: {mapping.pattern}\n"
    pretty += f"   Header Row Index: {mapping.header_row_index}\n"
    
    pretty += f"   Anchors ({len(mapping.anchors.columns)} columns):\n"
    for ref in mapping.anchors.columns:
        pretty += f"      - {ref.col_idx}: '{ref.detected_name}' → {ref.semantic_name} ({ref.data_type})\n"
    
    if mapping.pivot_axis:
        pretty += f"   Pivot Axis: {mapping.pivot_axis.col_range_start}:{mapping.pivot_axis.col_range_end}\n"
        pretty += f"      → New column: {mapping.pivot_axis.target_name} ({mapping.pivot_axis.extracted_type})\n"
    
    if mapping.value_body:
        pretty += f"   Value Body: {mapping.value_body.target_name} ({mapping.value_body.data_type})\n"
    
    if mapping.rows_to_exclude_keywords:
        pretty += f"   Rows to Exclude: {mapping.rows_to_exclude_keywords}\n"
    
    return pretty

def format_results_for_choregraph(results):
    """Format results for choregraph pipeline.

    Uses the human-readable ``label`` (slugified) as dict key when available,
    falling back to ``table_id``.  Deduplicates keys by appending ``_2``, ``_3``
    on collision.  Logs warnings for failed tables.
    """
    formatted_result = {}
    used_keys: Dict[str, int] = {}
    failed = []

    for result in results:
        if result['status'] == 'success':
            raw_key = result.get('label') or result.get('table_id', 'table')
            key = _slugify(raw_key)

            # Deduplicate
            if key in used_keys:
                used_keys[key] += 1
                key = f"{key}_{used_keys[key]}"
            else:
                used_keys[key] = 1

            formatted_result[key] = result['df']
        else:
            failed.append(result)
            logger.warning(
                "Table '%s' failed with status '%s': %s",
                result.get('table_id', '?'),
                result.get('status'),
                result.get('errors', result.get('message', 'unknown error'))
            )

    if failed:
        logger.warning(
            "%d/%d table(s) failed processing and were excluded from output",
            len(failed), len(results)
        )

    return formatted_result