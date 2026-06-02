# SPDX-FileCopyrightText: 2026 NVEIL SAS
# SPDX-FileContributor: Clément Baraille
# SPDX-FileContributor: Pierre Jacquet
# SPDX-License-Identifier: AGPL-3.0-or-later

import logging

import openpyxl
import pandas as pd
from openpyxl.utils import range_boundaries, get_column_letter
from typing import List, Optional, Dict, Any, Literal, Union
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


class SpreadsheetEncoder:
    """
    Encode un fichier Excel en représentation textuelle pour les LLMs.

    Pre-processes merged cells: captures merge info for visual markers,
    then unmerges and fills values so downstream DataFrames are clean.
    """

    def __init__(self, file_path, head_rows=10, tail_rows=5, head_cols=10, tail_cols=3,
                 show_empty_rows: bool = True, context_around_empty: int = 2):
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.head_rows = head_rows
        self.tail_rows = tail_rows
        self.head_cols = head_cols
        self.tail_cols = tail_cols
        self.show_empty_rows = show_empty_rows
        self.context_around_empty = context_around_empty

        # Load a second workbook with formulas preserved for metadata extraction
        self.formula_maps: Dict[str, Dict[tuple, str]] = {}
        try:
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            for sname in wb_formulas.sheetnames:
                fmap = {}
                ws = wb_formulas[sname]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            fmap[(cell.row, cell.column)] = cell.value
                self.formula_maps[sname] = fmap
            wb_formulas.close()
        except Exception:
            pass  # Graceful degradation (corrupted, .xls converted, etc.)

        # P2 fix: capture merge info BEFORE unmerging, for all sheets
        self.merge_maps: Dict[str, Dict] = {}
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            self.merge_maps[sheet_name] = self._capture_merged_cells(sheet)
            self.unmerge_and_fill_cells(sheet)

    def _capture_merged_cells(self, sheet) -> Dict:
        """Capture merge info from a sheet BEFORE unmerging. Returns a map of (row, col) -> merge info."""
        merged_map = {}
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_value = sheet.cell(row=min_row, column=min_col).value

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    is_origin = (r == min_row and c == min_col)
                    merged_map[(r, c)] = {"val": top_left_value, "origin": is_origin}
        return merged_map

    def unmerge_and_fill_cells(self, sheet):
        """
        Parcourt toutes les plages fusionnées, copie la valeur de la cellule
        en haut à gauche dans toutes les autres cellules, puis défusionne.
        Modifie la feuille en place.
        """
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            top_left_value = top_left_cell.value

            sheet.unmerge_cells(str(merged_range))

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

    def get_real_dimensions(self, sheet):
        """Trouve la dernière ligne et colonne contenant réellement des données."""
        if sheet.max_row == 0 or sheet.max_column == 0:
            return 0, 0

        real_max_row = 0
        for r in range(sheet.max_row, 0, -1):
            is_empty = all(sheet.cell(row=r, column=c).value is None for c in range(1, sheet.max_column + 1))
            if not is_empty:
                real_max_row = r
                break
        
        real_max_col = 0
        for c in range(sheet.max_column, 0, -1):
            is_empty = all(sheet.cell(row=r, column=c).value is None for r in range(1, sheet.max_row + 1))
            if not is_empty:
                real_max_col = c
                break
                
        return real_max_row, real_max_col
    
    def format_value(self, val):
        if val is None: return ""
        if isinstance(val, float): return f"{val:.2f}"
        if isinstance(val, str):
            val = val.replace("\n", " ").strip()
            return val[:50] + ".." if len(val) > 50 else val
        return str(val)

    def _is_row_empty(self, sheet, row, max_col):
        return all(sheet.cell(row=row, column=c).value is None for c in range(1, max_col + 1))

    def _find_empty_rows_in_range(self, sheet, start_row, end_row, max_col):
        empty_rows = []
        for r in range(start_row, end_row + 1):
            if self._is_row_empty(sheet, r, max_col):
                empty_rows.append(r)
        return empty_rows

    def _get_rows_to_display(self, total_rows, max_col, sheet):
        if total_rows <= self.head_rows + self.tail_rows:
            return list(range(1, total_rows + 1)), []
        
        rows_to_show = set(range(1, self.head_rows + 1))
        rows_to_show.update(range(total_rows - self.tail_rows + 1, total_rows + 1))
        
        omit_start = self.head_rows + 1
        omit_end = total_rows - self.tail_rows
        
        if self.show_empty_rows and omit_start <= omit_end:
            empty_rows = self._find_empty_rows_in_range(sheet, omit_start, omit_end, max_col)
            
            for empty_row in empty_rows:
                rows_to_show.add(empty_row)
                for offset in range(1, self.context_around_empty + 1):
                    ctx_row = empty_row - offset
                    if ctx_row >= 1:
                        rows_to_show.add(ctx_row)
                for offset in range(1, self.context_around_empty + 1):
                    ctx_row = empty_row + offset
                    if ctx_row <= total_rows:
                        rows_to_show.add(ctx_row)
        
        sorted_rows = sorted(rows_to_show)
        
        skip_zones = []
        for i in range(len(sorted_rows) - 1):
            if sorted_rows[i + 1] - sorted_rows[i] > 1:
                skip_zones.append((sorted_rows[i] + 1, sorted_rows[i + 1] - 1))
        
        return sorted_rows, skip_zones

    def _get_display_ranges(self, total, head, tail):
        if total <= head + tail:
            return list(range(1, total + 1)), None, None
        
        head_indices = list(range(1, head + 1))
        tail_indices = list(range(total - tail + 1, total + 1))
        skip_start = head + 1
        skip_end = total - tail
        
        return head_indices + tail_indices, skip_start, skip_end

    @staticmethod
    def _compact_ranges(numbers: List[int]) -> str:
        """Convert a sorted list of integers into compact range notation.

        Example: [1, 2, 3, 5, 7, 8, 9] → "1-3, 5, 7-9"
        """
        if not numbers:
            return ""
        nums = sorted(set(numbers))
        ranges = []
        start = end = nums[0]
        for n in nums[1:]:
            if n == end + 1:
                end = n
            else:
                ranges.append(f"{start}-{end}" if end > start else str(start))
                start = end = n
        ranges.append(f"{start}-{end}" if end > start else str(start))
        return ", ".join(ranges)

    def _get_row_flags(self, sheet_name: str, row: int, min_col: int, max_col: int) -> str:
        """Compute per-row formatting flags for inline display.

        Returns a compact string of flags:
          B = majority (>50%) of non-empty cells are bold
          Σ = majority (>50%) of non-empty cells contain formulas
          ■ = majority (>50%) of non-empty cells are shaded/colored
        Returns "" if no flags apply.
        """
        sheet = self.wb[sheet_name]
        fmap = self.formula_maps.get(sheet_name, {})

        non_empty = 0
        bold_count = 0
        formula_count = 0
        shaded_count = 0

        for c in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=c)
            if cell.value is None:
                continue
            non_empty += 1

            if (row, c) in fmap:
                formula_count += 1
            if cell.font and cell.font.bold is True:
                bold_count += 1
            if cell.fill and cell.fill.patternType not in (None, "none"):
                fg = cell.fill.fgColor
                if fg and fg.rgb and fg.rgb not in ("00000000", "FFFFFFFF"):
                    shaded_count += 1

        if non_empty == 0:
            return ""

        flags = ""
        if bold_count / non_empty > 0.5:
            flags += "B"
        if formula_count / non_empty > 0.5:
            flags += "\u03a3"
        if shaded_count / non_empty > 0.5:
            flags += "\u25a0"
        return flags

    def _build_column_metadata(self, sheet_name: str, min_row: int, max_row: int,
                               min_col: int, max_col: int) -> str:
        """Build a [COLUMN METADATA] summary for a rectangular region.

        Reports column-level patterns (formula/bold/shaded columns) and
        individual formula cells not covered by column patterns.
        Row-level patterns are handled by inline flags (⚑ column).
        Returns an empty string if no metadata is detected.
        """
        sheet = self.wb[sheet_name]
        fmap = self.formula_maps.get(sheet_name, {})

        formula_cells = set()
        non_empty_by_col: Dict[int, int] = {}

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                if cell.value is not None:
                    non_empty_by_col[c] = non_empty_by_col.get(c, 0) + 1
                if (r, c) in fmap:
                    formula_cells.add((r, c))

        if not formula_cells:
            return ""

        # --- Column-level formula patterns (>80% of non-empty data cells) ---
        formula_cols = []
        covered_by_col_formula = set()

        for c in range(min_col, max_col + 1):
            total = non_empty_by_col.get(c, 0)
            if total == 0:
                continue
            col_formulas = {(r, c) for r in range(min_row, max_row + 1) if (r, c) in formula_cells}
            if len(col_formulas) / total > 0.8:
                rows_with = sorted(r for r, _ in col_formulas)
                formula_cols.append((get_column_letter(c), self._compact_ranges(rows_with)))
                covered_by_col_formula.update(col_formulas)

        # --- Individual formula cells (not covered by column patterns) ---
        remaining_formulas = formula_cells - covered_by_col_formula
        individual_cells = [f"{get_column_letter(c)}{r}" for r, c in sorted(remaining_formulas)]

        # --- Build output ---
        lines = ["[COLUMN METADATA]"]
        has_content = False

        if formula_cols:
            parts = [f"{letter} (rows {rows})" for letter, rows in formula_cols]
            lines.append(f"Formula columns: {', '.join(parts)}")
            has_content = True

        if individual_cells:
            lines.append(f"Individual formula cells: {', '.join(individual_cells)}")
            has_content = True

        return "\n".join(lines) if has_content else ""

    def encode_sheet(self, sheet_name):
        sheet = self.wb[sheet_name]

        # Use pre-captured merge map (captured before unmerging in __init__)
        merged_map = self.merge_maps.get(sheet_name, {})

        max_r, max_c = self.get_real_dimensions(sheet)
        
        output = [f"## SHEET: {sheet_name} (Dimensions: {max_r} rows × {max_c} cols)"]
        
        row_indices, skip_zones = self._get_rows_to_display(max_r, max_c, sheet)
        col_indices, col_skip_start, col_skip_end = self._get_display_ranges(max_c, self.head_cols, self.tail_cols)
        
        # Header (with inline flags column ⚑)
        header_parts = ["| | ⚑ |"]
        for idx, c in enumerate(col_indices):
            if col_skip_start and idx == self.head_cols:
                header_parts.append(f" ... |")
            header_parts.append(f" {get_column_letter(c)} |")
        header = "".join(header_parts)

        sep_count = len(col_indices) + (1 if col_skip_start else 0)
        separator = "|---|---|" + "---|" * sep_count
        
        output.extend([header, separator])
        
        prev_row = 0

        for r in row_indices:
            if prev_row > 0 and r - prev_row > 1:
                skip_start = prev_row + 1
                skip_end = r - 1
                skip_count = skip_end - skip_start + 1
                skip_msg = f"| ... | ⋮ ({skip_count} lignes {skip_start}→{skip_end} omises) |"
                output.append(skip_msg)
            
            prev_row = r
            
            row_cells = []
            row_is_empty = True
            
            for col_idx, c in enumerate(col_indices):
                if col_skip_start and col_idx == self.head_cols:
                    row_cells.append("...")
                
                cell_val = ""

                if (r, c) in merged_map:
                    m_info = merged_map[(r, c)]
                    cell_val = f"{self.format_value(m_info['val'])} (M)" if m_info['origin'] else "〃"
                    row_is_empty = False
                else:
                    val = sheet.cell(row=r, column=c).value
                    if val is not None:
                        cell_val = self.format_value(val)
                        row_is_empty = False
                
                row_cells.append(cell_val)

            if row_is_empty:
                output.append(f"| {r} | | ── (empty row) ── |")
            else:
                flags = self._get_row_flags(sheet_name, r, 1, max_c)
                output.append(f"| {r} | {flags} | " + " | ".join(row_cells) + " |")

        # Append column metadata summary (formulas)
        metadata = self._build_column_metadata(sheet_name, 1, max_r, 1, max_c)
        if metadata:
            output.append("")
            output.append(metadata)

        return "\n".join(output)

    def encode_all_sheets(self):
        return "\n\n" + "\n\n".join([self.encode_sheet(name) for name in self.wb.sheetnames])

    def get_dataframes(self) -> Dict[str, pd.DataFrame]:
        """Build DataFrames from the openpyxl workbook (already unmerged and filled).

        This provides a single source of truth: the LLM encoding and the
        DataFrames used by the ETL engine both come from the same openpyxl
        data, eliminating value disagreements (P3 fix).
        """
        result = {}
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            max_r, max_c = self.get_real_dimensions(sheet)
            if max_r == 0 or max_c == 0:
                result[sheet_name] = pd.DataFrame()
                continue

            data = []
            for row in sheet.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True):
                data.append(list(row))

            df = pd.DataFrame(data)
            df.columns = [get_column_letter(i + 1) for i in range(df.shape[1])]
            result[sheet_name] = df
        return result
    
    def render_range_as_markdown(self, sheet_name: str, excel_range: str, max_rows: int = 15) -> str:
        """
        Génère le rendu visuel riche (avec merges) pour une plage spécifique.
        Remplace df_to_markdown pour le contexte LLM.
        """
        sheet = self.wb[sheet_name]
        merged_map = self.merge_maps.get(sheet_name, {})
        
        # Récupérer les bornes (1-based pour openpyxl)
        min_col, min_row, max_col, max_row = range_boundaries(excel_range)
        
        output = []

        # 1. Header: row number + flags + column letters
        header_parts = ["| | ⚑ |"]
        for c in range(min_col, max_col + 1):
            header_parts.append(f" {get_column_letter(c)} |")
        output.append("".join(header_parts))

        # 2. Séparateur
        output.append("|---|---|" + "---|" * (max_col - min_col + 1))
        
        # 3. Données
        rows_to_render = []
        
        # Logique de troncature si trop de lignes
        total_rows_in_range = max_row - min_row + 1
        if total_rows_in_range > max_rows:
            head_limit = min_row + max_rows - 3
            tail_start = max_row - 2
            
            rows_to_render.extend(range(min_row, head_limit + 1))
            rows_to_render.append(None) # Marqueur de saut
            rows_to_render.extend(range(tail_start, max_row + 1))
        else:
            rows_to_render.extend(range(min_row, max_row + 1))
            
        for r in rows_to_render:
            if r is None:
                output.append(f"| ... | | ⋮ ({total_rows_in_range - max_rows} lignes omises) |")
                continue

            row_cells = []
            for c in range(min_col, max_col + 1):
                cell_val = ""
                if (r, c) in merged_map:
                    m_info = merged_map[(r, c)]
                    cell_val = f"{self.format_value(m_info['val'])} (M)" if m_info['origin'] else "〃"
                else:
                    val = sheet.cell(row=r, column=c).value
                    if val is not None:
                        cell_val = self.format_value(val)

                row_cells.append(cell_val)

            flags = self._get_row_flags(sheet_name, r, min_col, max_col)
            output.append(f"| {r} | {flags} | " + " | ".join(row_cells) + " |")

        # Append column metadata summary (formulas)
        metadata = self._build_column_metadata(sheet_name, min_row, max_row, min_col, max_col)
        if metadata:
            output.append("")
            output.append(metadata)

        return "\n".join(output)
